import streamlit as st
import re
from io import BytesIO

# DOCX
from docx import Document
from docx.shared import RGBColor

# Excel
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# PDF
from PyPDF2 import PdfReader, PdfWriter
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter

st.set_page_config(page_title="Universal Proofreader", layout="wide")
st.title("📝 Multilingual Proofreader - Highlight Foreign Text")

# Sidebar: Settings
st.sidebar.header("Settings")
source_lang = st.sidebar.selectbox("Source Language", ["English", "Hindi", "Kannada"])
dest_lang = st.sidebar.selectbox("Target Language", ["Hindi", "Kannada"])

# Unicode ranges for scripts
unicode_ranges = {
    "Hindi": r"\u0900-\u097F",
    "Kannada": r"\u0C80-\u0CFF",
}

# File uploader
uploaded_file = st.file_uploader(
    "📄 Upload your document",
    type=["txt", "docx", "pdf", "xls", "xlsx", "csv"]
)

# --- Helper functions ---

def detect_foreign_text(text, dest_lang):
    """Return a list of foreign words not in the target language script."""
    dest_range = unicode_ranges.get(dest_lang, "")
    pattern = rf"([^{dest_range}\s\d.,;:!?\"'()\-\n]+)"
    return re.findall(pattern, text)

# --- DOCX processing ---
def process_docx(file):
    doc = Document(file)
    dest_range = unicode_ranges.get(dest_lang, "")
    pattern = re.compile(rf"([^{dest_range}\s\d.,;:!?\"'()\-\n]+)")
    for para in doc.paragraphs:
        for run in para.runs:
            if re.search(pattern, run.text):
                run.font.highlight_color = 7  # Yellow
                run.font.color.rgb = RGBColor(0,0,0)
    out = BytesIO()
    doc.save(out)
    return out

# --- Excel processing ---
def process_excel(file):
    wb = load_workbook(filename=BytesIO(file.read()))
    fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    dest_range = unicode_ranges.get(dest_lang, "")
    pattern = re.compile(rf"([^{dest_range}\s\d.,;:!?\"'()\-\n]+)")
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    if re.search(pattern, cell.value):
                        cell.fill = fill
    out = BytesIO()
    wb.save(out)
    return out

# --- CSV / Excel CSV processing ---
def process_csv(file):
    df = pd.read_csv(file)
    dest_range = unicode_ranges.get(dest_lang, "")
    pattern = re.compile(rf"([^{dest_range}\s\d.,;:!?\"'()\-\n]+)")
    # Create a new column to mark foreign text
    df['Foreign Text Detected'] = df.apply(lambda row: ', '.join([str(v) for v in row if re.search(pattern, str(v))]), axis=1)
    out = BytesIO()
    df.to_csv(out, index=False)
    return out

# --- TXT processing ---
def process_txt(file):
    text = file.read().decode('utf-8', errors='ignore')
    foreign_words = detect_foreign_text(text, dest_lang)
    highlighted_text = text
    for word in set(foreign_words):
        highlighted_text = highlighted_text.replace(word, f"**{word}**")
    return highlighted_text.encode('utf-8')

# --- PDF processing ---
def process_pdf(file):
    # Extract text
    reader = PdfReader(file)
    out_pdf = PdfWriter()
    for page in reader.pages:
        out_pdf.add_page(page)
    # For now, PDF highlighting is not inline; users can view HTML preview or export DOCX
    out = BytesIO()
    out_pdf.write(out)
    return out

# --- Main ---
if uploaded_file:
    ext = uploaded_file.name.split('.')[-1].lower()
    st.subheader("🔍 Proofreading Preview")

    if ext == "docx":
        output = process_docx(uploaded_file)
        st.success("✅ DOCX processed with highlights!")
        st.download_button("⬇️ Download DOCX", data=output.getvalue(),
                           file_name=f"proofread_{dest_lang}.docx",
                           mime="application/vnd.openxmlformats-officedocument.wordprocessingml.document")
    elif ext in ["xls", "xlsx"]:
        output = process_excel(uploaded_file)
        st.success("✅ Excel processed with highlights!")
        st.download_button("⬇️ Download Excel", data=output.getvalue(),
                           file_name=f"proofread_{dest_lang}.{ext}",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    elif ext == "csv":
        output = process_csv(uploaded_file)
        st.success("✅ CSV processed with foreign text column!")
        st.download_button("⬇️ Download CSV", data=output.getvalue(),
                           file_name=f"proofread_{dest_lang}.csv",
                           mime="text/csv")
    elif ext == "txt":
        output = process_txt(uploaded_file)
        st.success("✅ TXT processed with markers for foreign text!")
        st.download_button("⬇️ Download TXT", data=output,
                           file_name=f"proofread_{dest_lang}.txt",
                           mime="text/plain")
    elif ext == "pdf":
        st.warning("⚠ PDF highlighting not inline. Downloaded PDF same as original. Use DOCX for highlights.")
        output = process_pdf(uploaded_file)
        st.download_button("⬇️ Download PDF", data=output.getvalue(),
                           file_name=f"proofread_{dest_lang}.pdf",
                           mime="application/pdf")
    else:
        st.error("❌ Unsupported file type")
else:
    st.info("Upload a document to start proofing.")
